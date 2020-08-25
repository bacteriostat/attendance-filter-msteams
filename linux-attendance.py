# Imports

import os
from openpyxl import Workbook
from openpyxl import load_workbook
import re
import datetime
import csv
import sys

#Lists for storing different parsed data

late = []
onTime = []
wholeNoisy = []
whole = []

#File initializers and check for invalid files
if re.search("^.+([.]csv)$" , sys.argv[1]):
    attendance_csv = sys.argv[1]
else:
    sys.exit("Invalid attendance file. The file should be in csv format.")

if re.search("^.+([.]xlsx)$", sys.argv[2]):
    student_list = sys.argv[2]
else:
    sys.exit("Invalid student data file. The file must be in xlsx format.")

output_location = os.path.join(os.getcwd(), "Attendance-"+str(datetime.datetime.now()))
os.mkdir(output_location)

#parsing attendance file (csv)

with open(attendance_csv, newline='', encoding = "UTF-16") as attendance_file:
    attendance_log = csv.reader(attendance_file, delimiter=' ', quotechar='|')
    for row in attendance_log:
      parsedRow = ' '.join(row).split("\t")  
      wholeNoisy.append(parsedRow)

#Remove and store header from whole list

wholeNoisy.pop(0)

#Remove irrelevant data

for i in range(0, len(wholeNoisy)):
    entry = wholeNoisy[i]
    if entry == -1:
        continue
    if entry[1]=='Left':
        #didJoinBack = False
        temp = wholeNoisy[i]
        wholeNoisy[i]=-1
        for j in range(i+1, len(wholeNoisy)):
            if wholeNoisy[j][0] == temp[0]:
                wholeNoisy[j]=-1
    else:
        whole.append(entry)

#Filter out late comers

for row in whole:
    if re.search("\d{2}:\d{2}:\d{2}", row[2]):
        if int(row[2].split(':')[1])>30:
            late.append(row)
        else:
            onTime.append(row) 

#Read student data 

student_list_wb = load_workbook(filename = student_list, read_only=True)
student_list_sheet = student_list_wb["Sheet1"]

#Create workbook

late_wb = Workbook()
absentees_wb = Workbook()

#Header for output files
late_header = ["Name", "User Action", "Timestamp", "Class & Section", "Enrollment No.", "Linux Install?", "Lab Performance"]
absentees_header = ["Section", "Roll No.", "Student Name", "Enrollment No.", "Linux Install?", "Lab Performance"]

#Create worksheet
late_ws = late_wb.active
absentees_ws = absentees_wb.active

#Write to files
late_ws.append(late_header)
absentees_ws.append(absentees_header)

for read_only_data in student_list_sheet.iter_rows(min_row = 2, max_col = 7, values_only=True):
    data = list(read_only_data)

    present = False
    
    for late_comer in late:
        if data[3].strip() == late_comer[0].strip():
            late_ws.append(late_comer + data[1:2] + data[4:])
            present=True
    
    for punctual in onTime:
        if data[3].strip() == punctual[0].strip():
            present=True

    if not present:
        absentees_ws.append(data[1:]) 



#Save files
late_wb.save(os.path.join(output_location, "late.xlsx"))
absentees_wb.save(os.path.join(output_location, "absentees.xlsx"))

print("Saved output files in "+output_location)

